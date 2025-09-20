import openpyxl as xl
wb=xl.load_workbook('Transaction.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)

for row in range(1, sheet.max_row+1):
    print(row)

#print value
for row in range(1, sheet.max_row+1):
    cell1=sheet.cell(row, 1)
    cell2=sheet.cell(row, 2)
    cell3=sheet.cell(row, 3)
    cell4=sheet.cell(row, 4)
    print(cell1.value," ", cell2.value," ", cell3.value," ", cell4.value)

#update price
for row in range(2, sheet.max_row+1):
    cell=sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('Transaction2.xlsx')


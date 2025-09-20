import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# Load workbook from full path

def proccess_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Print row numbers
    for row in range(1, sheet.max_row + 1):
        print(row)

    # Print values
    for row in range(1, sheet.max_row + 1):
        cell1 = sheet.cell(row, 1)
        cell2 = sheet.cell(row, 2)
        cell3 = sheet.cell(row, 3)
        cell4 = sheet.cell(row, 4)
        print(cell1.value, " ", cell2.value, " ", cell3.value, " ", cell4.value)

    # Update price
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price=cell.value*0.9
        corrected_price_cell=sheet.cell(row, 4)
        corrected_price_cell.value=corrected_price

    values=Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Save back to same file
    wb.save(filename)


#After run this type ".\Transaction.xlsx" in th terminal
